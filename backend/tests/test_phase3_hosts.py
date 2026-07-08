"""Phase 3: Hosts 高级功能测试"""
from django.test import TestCase
from django.urls import reverse
from rest_framework import status
from rest_framework.test import APIClient
from rest_framework_simplejwt.tokens import RefreshToken
from backend.models.user import Users
from backend.models.host import Host, Cluster
from backend.models.authority import Authority, UserAuthority
from backend.services.host import HostService
from unittest.mock import patch, MagicMock
import io


class HostServiceAdvancedTest(TestCase):
    def test_generate_random_password(self):
        pwd = HostService.generate_random_password(16)
        self.assertEqual(len(pwd), 16)

    def test_hash_password(self):
        h = HostService.hash_password("test", "key")
        self.assertEqual(len(h), 32)

    def test_batch_update_password(self):
        c = Cluster.objects.create(name="c1")
        h1 = Host.objects.create(hostname="h1", ip_address="1.1.1.1", username="root", password="old", cluster=c)
        h2 = Host.objects.create(hostname="h2", ip_address="1.1.1.2", username="root", password="old", cluster=c)
        result = HostService.batch_update_password([h1.id, h2.id], "newpass123", "key")
        self.assertEqual(result["updated"], 2)
        self.assertEqual(len(result["failed"]), 0)
        h1.refresh_from_db()
        self.assertEqual(h1.password, HostService.hash_password("newpass123", "key"))

    def test_batch_update_password_nonexistent(self):
        result = HostService.batch_update_password([99999], "newpass")
        self.assertEqual(result["updated"], 0)
        self.assertEqual(len(result["failed"]), 1)

    @patch("backend.utils.ssh.SSHClient")
    def test_remote_command(self, mock_ssh_class):
        mock_client = MagicMock()
        mock_client.__enter__ = MagicMock(return_value=mock_client)
        mock_client.__exit__ = MagicMock(return_value=False)
        mock_client.execute_command.return_value = ("stdout", "", 0)
        mock_ssh_class.return_value = mock_client

        c = Cluster.objects.create(name="c1")
        h = Host.objects.create(hostname="h1", ip_address="1.1.1.1", username="root", password="pass", port=22, cluster=c)
        result = HostService.remote_command(h.id, "ls -l")
        self.assertEqual(result["exit_code"], 0)
        self.assertEqual(result["stdout"], "stdout")

    def test_export_hosts_to_excel(self):
        c = Cluster.objects.create(name="c1")
        Host.objects.create(hostname="h1", ip_address="1.1.1.1", username="root", cluster=c)
        data = HostService.export_hosts_to_excel()
        self.assertIsInstance(data, bytes)
        self.assertTrue(len(data) > 0)

    def test_import_hosts_from_excel(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["hostname", "ip_address", "port", "username", "status", "os_type", "host_type"])
        ws.append(["h1", "1.1.1.1", 22, "root", "online", "centos", "VMHost"])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        result = HostService.import_hosts_from_excel(buffer)
        self.assertEqual(result["created"], 1)
        self.assertEqual(len(result["errors"]), 0)
        self.assertTrue(Host.objects.filter(ip_address="1.1.1.1").exists())

    def test_import_key_cloud(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([
            "序列号", "IP地址", "主机名", "ntp地址", "密码", "设备分类",
            "是否有集群属性", "是否有专区属性", "是否绑定cell",
            "带外vlan", "带内管理vlan", "带内管理接口名称",
            "存储vlan", "存储网络接口名称",
            "业务网络vlan", "业务网络接口名称",
            "其他网络vlan", "其他网络接口名称",
            "RAID要求", "BIOS配置要求",
            "双引擎系统盘分区要求(单位为M)", "单引擎系统盘分区要求(单位为M)",
            "双引擎推荐操作系统版本", "单引擎操作系统",
            "存储IP", "业务IP", "其他IP",
        ])
        ws.append([
            "SN001", "10.0.0.1", "host-01", "ntp.example.com", "pass123", "server",
            "是", "否", "是",
            "100", "200", "eth0",
            "300", "eth1",
            "400", "eth2",
            "500", "eth3",
            "RAID1", "UEFI",
            "1024", "512",
            "CentOS8", "CentOS7",
            "192.168.1.1", "192.168.2.1", "192.168.3.1",
        ])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        result = HostService.import_key_cloud(buffer)
        self.assertEqual(result["created"], 1)
        self.assertEqual(result["updated"], 0)

        host = Host.objects.get(hostname="host-01")
        self.assertEqual(host.serial_number, "SN001")
        self.assertEqual(host.ip_address, "10.0.0.1")
        self.assertEqual(host.ntp_address, "ntp.example.com")
        self.assertEqual(host.password, "pass123")
        self.assertEqual(host.host_type, "server")
        self.assertTrue(host.is_cluster_type)
        self.assertFalse(host.is_zone_type)
        self.assertTrue(host.is_bind_cell_type)
        self.assertEqual(host.ipmi_vlan, "100")
        self.assertEqual(host.manage_vlan, "200")
        self.assertEqual(host.manage_nic1, "eth0")
        self.assertEqual(host.storage_vlan, "300")
        self.assertEqual(host.storage_ifname, "eth1")
        self.assertEqual(host.business_vlan, "400")
        self.assertEqual(host.business_ifname, "eth2")
        self.assertEqual(host.other_vlan, "500")
        self.assertEqual(host.other_ifname, "eth3")
        self.assertEqual(host.raid, "RAID1")
        self.assertEqual(host.bios_config, "UEFI")
        self.assertEqual(host.mount_info, "512")  # 单引擎覆盖双引擎
        self.assertEqual(host.os_version, "CentOS7")  # 单引擎覆盖双引擎
        self.assertEqual(str(host.storage_address), "192.168.1.1")
        self.assertEqual(str(host.business_address), "192.168.2.1")
        self.assertEqual(str(host.other_address), "192.168.3.1")

    def test_import_key_cloud_update_existing(self):
        import openpyxl
        Host.objects.create(hostname="host-02", ip_address="10.0.0.2", username="root")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["主机名", "IP地址", "密码"])
        ws.append(["host-02", "10.0.0.2", "newpass"])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        result = HostService.import_key_cloud(buffer)
        self.assertEqual(result["created"], 0)
        self.assertEqual(result["updated"], 1)

        host = Host.objects.get(hostname="host-02")
        self.assertEqual(host.password, "newpass")


class HostViewAdvancedTest(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = Users.objects.create(
            user="testuser",
            nickname="Test User",
            email="test@example.com",
            enable=1,
        )
        self.user.set_password("testpass123")
        self.user.save()
        self.admin_auth = Authority.objects.create(authority_id=888, authority_name="超级管理员")
        UserAuthority.objects.create(user=self.user, authority=self.admin_auth)
        refresh = RefreshToken.for_user(self.user)
        self.access_token = str(refresh.access_token)
        self.client.credentials(HTTP_AUTHORIZATION=f"Bearer {self.access_token}")
        self.cluster = Cluster.objects.create(name="c1")
        self.host = Host.objects.create(
            hostname="h1", ip_address="1.1.1.1", username="root",
            password="pass", port=22, cluster=self.cluster
        )

    @patch("backend.services.host.HostService.batch_update_password")
    def test_batch_update_password_action(self, mock_batch):
        mock_batch.return_value = {"success": True, "updated": 1, "message": "ok", "password": "new", "failed": []}
        url = reverse("host-batch-update-password")
        response = self.client.post(url, {"host_ids": [self.host.id], "password": "newpass"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["errno"], 0)

    @patch("backend.services.host.HostService.remote_command")
    def test_remote_command_action(self, mock_cmd):
        mock_cmd.return_value = {"success": True, "message": "ok", "stdout": "out", "stderr": "", "exit_code": 0}
        url = reverse("host-remote-command", args=[self.host.id])
        response = self.client.post(url, {"command": "ls -l"}, format="json")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["errno"], 0)

    def test_import_key_cloud_action(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["主机名", "IP地址", "密码"])
        ws.append(["cloud-host", "10.0.0.99", "cloudpass"])
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        url = reverse("host-import-key-cloud")
        response = self.client.post(url, {"file": buffer}, format="multipart")
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["errno"], 0)
        self.assertTrue(Host.objects.filter(hostname="cloud-host").exists())
