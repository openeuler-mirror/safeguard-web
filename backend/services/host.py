"""集群相关服务"""
import random
import string
import hashlib
import logging
from dataclasses import dataclass
from typing import Optional, Dict, List
from backend.models.host import Cluster, Host, VM


@dataclass
class VMOperationResult:
    """VM 操作结果"""
    success: bool
    vm: Optional[VM] = None
    message: str = ""
    is_simulation: bool = False
    new_status: Optional[str] = None
from backend.utils.hardware_collector import (
    collect_host_hardware,
    collect_host_lldp,
    update_host_hardware_info,
    update_host_lldp_info,
    collect_all_hardware_info,
)
from backend.common.exceptions import (
    HostNotFoundError,
    VMNotFoundError,
    HardwareCollectError,
    LLDCollectError,
    PasswordUpdateError,
    HostImportError,
    RemoteCommandError,
    VMOperationError,
    OperationError,
)

logger = logging.getLogger(__name__)

# libvirt 状态映射
LIBVIRT_STATE_MAP = {
    0: 'nostate',      # VIR_DOMAIN_NOSTATE
    1: 'running',      # VIR_DOMAIN_RUNNING
    2: 'blocked',      # VIR_DOMAIN_BLOCKED
    3: 'paused',        # VIR_DOMAIN_PAUSED
    4: 'shutdown',      # VIR_DOMAIN_SHUTDOWN
    5: 'shutoff',       # VIR_DOMAIN_SHUTOFF
    6: 'crashed',       # VIR_DOMAIN_CRASHED
    7: 'suspended',     # VIR_DOMAIN_PMSUSPENDED
}


class ClusterService:
    """集群服务"""

    @staticmethod
    def list_clusters(filters: Optional[dict] = None, page: int = 1, page_size: int = 10):
        """获取集群列表（支持分页和过滤）"""
        queryset = Cluster.objects.all()
        if filters:
            queryset = queryset.filter(**filters)

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = list(queryset[start:end])

        return {
            'total': total,
            'page': page,
            'page_size': page_size,
            'results': results
        }

    @staticmethod
    def get_cluster(cluster_id: int) -> Optional[Cluster]:
        """获取集群详情"""
        try:
            return Cluster.objects.get(pk=cluster_id)
        except Cluster.DoesNotExist:
            return None

    @staticmethod
    def create_cluster(data: dict) -> Cluster:
        """创建集群"""
        return Cluster.objects.create(**data)

    @staticmethod
    def update_cluster(cluster_id: int, data: dict) -> Optional[Cluster]:
        """更新集群"""
        try:
            cluster = Cluster.objects.get(pk=cluster_id)
            for key, value in data.items():
                setattr(cluster, key, value)
            cluster.save()
            return cluster
        except Cluster.DoesNotExist:
            return None

    @staticmethod
    def delete_cluster(cluster_id: int) -> bool:
        """删除集群"""
        try:
            cluster = Cluster.objects.get(pk=cluster_id)
            cluster.delete()
            return True
        except Cluster.DoesNotExist:
            return False

    @staticmethod
    def get_cluster_topology(cluster_id: int) -> Optional[Dict]:
        """
        获取集群拓扑

        返回集群下所有主机和VM的关联关系

        Returns:
            {
                'cluster': {...},
                'hosts': [...],
                'vms': [...],
                'lldp_topology': {...}  # 如果有LLDP信息
            }
        """
        try:
            cluster = Cluster.objects.get(pk=cluster_id)
        except Cluster.DoesNotExist:
            return None

        hosts = list(cluster.host_set.all())
        vms = list(VM.objects.filter(host__in=hosts).select_related('host', 'cluster'))

        # 构建 LLDP 拓扑信息
        lldp_topology = {}
        for host in hosts:
            if host.lldp_infos:
                lldp_topology[host.ip_address] = host.lldp_infos

        return {
            'cluster': {
                'id': cluster.id,
                'name': cluster.name,
                'description': cluster.description,
            },
            'hosts': [
                {
                    'id': h.id,
                    'hostname': h.hostname,
                    'ip_address': h.ip_address,
                    'status': h.status,
                    'lldp_count': len(h.lldp_infos) if h.lldp_infos else 0,
                }
                for h in hosts
            ],
            'vms': [
                {
                    'id': v.id,
                    'name': v.name,
                    'uuid': v.uuid,
                    'status': v.status,
                    'host_id': v.host_id,
                    'ip_address': v.ip_address,
                }
                for v in vms
            ],
            'lldp_topology': lldp_topology,
        }


class HostService:
    """主机服务"""

    @staticmethod
    def list_hosts(filters: Optional[dict] = None, page: int = 1, page_size: int = 10):
        """获取主机列表（支持分页和过滤）"""
        queryset = Host.objects.select_related('cluster').all()
        if filters:
            queryset = queryset.filter(**filters)

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = list(queryset[start:end])

        return {
            'total': total,
            'page': page,
            'page_size': page_size,
            'results': results
        }

    @staticmethod
    def get_host(host_id: int) -> Optional[Host]:
        """获取主机详情"""
        try:
            return Host.objects.select_related('cluster').get(pk=host_id)
        except Host.DoesNotExist:
            return None

    @staticmethod
    def create_host(data: dict) -> Host:
        """创建主机"""
        return Host.objects.create(**data)

    @staticmethod
    def update_host(host_id: int, data: dict) -> Optional[Host]:
        """更新主机"""
        try:
            host = Host.objects.get(pk=host_id)
            for key, value in data.items():
                setattr(host, key, value)
            host.save()
            return host
        except Host.DoesNotExist:
            return None

    @staticmethod
    def delete_host(host_id: int) -> bool:
        """删除主机"""
        try:
            host = Host.objects.get(pk=host_id)
            host.delete()
            return True
        except Host.DoesNotExist:
            return False

    @staticmethod
    def collect_hardware(host_id: int) -> Dict:
        """
        采集主机硬件信息

        Args:
            host_id: 主机ID

        Returns:
            {
                'arch_info': {...},
                'uptime': {...},
                'os_version': str,
                'cpu_info': {...},
                'disk_info': {...},
                'memory_info': {...},
                'network_info': {...}
            }

        Raises:
            HostNotFoundError: 主机不存在
            HardwareCollectError: 硬件信息采集失败
        """
        try:
            host = Host.objects.get(pk=host_id)
        except Host.DoesNotExist:
            logger.error(f"Host {host_id} not found")
            raise HostNotFoundError(host_id)

        success = update_host_hardware_info(host)
        if success:
            host.refresh_from_db()
            logger.info(f"Hardware info collected for host {host_id}")
            return {
                'arch_info': host.arch_info,
                'uptime': host.uptime,
                'os_version': host.os_version,
                'cpu_info': host.cpu_info,
                'disk_info': host.disk_info,
                'memory_info': host.memory_info,
                'network_info': host.network_info,
            }
        else:
            logger.error(f"Failed to collect hardware info for host {host_id}")
            raise HardwareCollectError('硬件信息采集失败')

    @staticmethod
    def collect_lldp(host_id: int) -> List:
        """
        采集 LLDP 拓扑信息

        Args:
            host_id: 主机ID

        Returns:
            LLDP 信息列表

        Raises:
            HostNotFoundError: 主机不存在
            LLDCollectError: LLDP 信息采集失败
        """
        try:
            host = Host.objects.get(pk=host_id)
        except Host.DoesNotExist:
            logger.error(f"Host {host_id} not found")
            raise HostNotFoundError(host_id)

        success = update_host_lldp_info(host)
        if success:
            host.refresh_from_db()
            logger.info(f"LLDP info collected for host {host_id}")
            return host.lldp_infos
        else:
            logger.error(f"Failed to collect LLDP info for host {host_id}")
            raise LLDCollectError('LLDP 信息采集失败')

    @staticmethod
    def get_host_lldp(host_id: int) -> List:
        """
        获取主机已保存的 LLDP 信息

        Args:
            host_id: 主机ID

        Returns:
            LLDP 信息列表

        Raises:
            HostNotFoundError: 主机不存在
        """
        try:
            host = Host.objects.get(pk=host_id)
        except Host.DoesNotExist:
            logger.error(f"Host {host_id} not found")
            raise HostNotFoundError(host_id)

        return host.lldp_infos or []

    @staticmethod
    def collect_all(host_id: int) -> Dict:
        """
        采集主机所有硬件信息（包括 LLDP）

        Args:
            host_id: 主机ID

        Returns:
            {
                'hardware': {
                    'arch_info': {...},
                    'uptime': {...},
                    'os_version': str,
                    'cpu_info': {...},
                    'disk_info': {...},
                    'memory_info': {...},
                    'network_info': {...},
                    'mount_info': {...},
                    'dmesg_info': {...}
                },
                'lldp': [...]
            }

        Raises:
            HostNotFoundError: 主机不存在
            HardwareCollectError: 信息采集失败
        """
        try:
            host = Host.objects.get(pk=host_id)
        except Host.DoesNotExist:
            logger.error(f"Host {host_id} not found")
            raise HostNotFoundError(host_id)

        try:
            update_host_hardware_info(host)
            update_host_lldp_info(host)

            host.refresh_from_db()
            logger.info(f"All info collected for host {host_id}")
            return {
                'hardware': {
                    'arch_info': host.arch_info,
                    'uptime': host.uptime,
                    'os_version': host.os_version,
                    'cpu_info': host.cpu_info,
                    'disk_info': host.disk_info,
                    'memory_info': host.memory_info,
                    'network_info': host.network_info,
                    'mount_info': host.mount_info,
                    'dmesg_info': host.dmesg_info,
                },
                'lldp': host.lldp_infos
            }
        except Exception as e:
            logger.error(f"Failed to collect all info for host {host_id}: {e}")
            raise HardwareCollectError(str(e))

    @staticmethod
    def generate_random_password(length: int = 16) -> str:
        """
        生成随机密码

        Args:
            length: 密码长度

        Returns:
            随机密码字符串
        """
        charset = string.ascii_letters + string.digits + "!@#$%&*"
        return ''.join(random.choice(charset) for _ in range(length))

    @staticmethod
    def hash_password(password: str, key: str = "culinux") -> str:
        """
        哈希密码

        Args:
            password: 明文密码
            key: 加密密钥

        Returns:
            哈希后的密码
        """
        hasher = hashlib.md5()
        hasher.update((password + key).encode('utf-8'))
        return hasher.hexdigest()

    @staticmethod
    def update_host_password(host_id: int, new_password: str = None, key: str = "culinux") -> str:
        """
        修改主机密码

        Args:
            host_id: 主机ID
            new_password: 新密码（如果为空则生成随机密码）
            key: 加密密钥

        Returns:
            新密码（明文）

        Raises:
            HostNotFoundError: 主机不存在
            PasswordUpdateError: 密码更新失败
        """
        try:
            host = Host.objects.get(pk=host_id)
        except Host.DoesNotExist:
            logger.error(f"Host {host_id} not found")
            raise HostNotFoundError(host_id)

        try:
            if not new_password:
                new_password = HostService.generate_random_password()

            encrypted_password = HostService.hash_password(new_password, key)

            host.password = encrypted_password
            host.save()

            logger.info(f"Password updated for host {host_id}")
            return new_password
        except Exception as e:
            logger.error(f"Failed to update password for host {host_id}: {e}")
            raise PasswordUpdateError(str(e))

    @staticmethod
    def import_hosts_from_excel(file_obj) -> Dict:
        """从 Excel 导入主机

        Returns:
            {
                "created": int,
                "updated": int,
                "errors": list
            }

        Raises:
            HostImportError: 导入失败
        """
        import openpyxl
        from io import BytesIO
        try:
            wb = openpyxl.load_workbook(BytesIO(file_obj.read()))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            created_count = 0
            updated_count = 0
            errors = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    data = dict(zip(headers, row))
                    if not data.get("hostname") or not data.get("ip_address"):
                        continue
                    host, created = Host.objects.update_or_create(
                        ip_address=data["ip_address"],
                        defaults={
                            "hostname": data.get("hostname", ""),
                            "port": data.get("port", 22),
                            "username": data.get("username", ""),
                            "password": data.get("password", ""),
                            "status": data.get("status", "offline"),
                            "os_type": data.get("os_type", ""),
                            "host_type": data.get("host_type", "VMHost"),
                            "cluster_id": data.get("cluster_id") or None,
                        }
                    )
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                except Exception as e:
                    errors.append(str(e))
            logger.info(f"Import completed: created {created_count}, updated {updated_count}, errors {len(errors)}")
            return {
                "created": created_count,
                "updated": updated_count,
                "errors": errors
            }
        except Exception as e:
            logger.error(f"Import hosts from excel failed: {e}")
            raise HostImportError(str(e))

    @staticmethod
    def import_key_cloud(file_obj) -> Dict:
        """从骨干云部署表 Excel 导入主机

        支持的中文字段映射：
        序列号、IP地址、主机名、ntp地址、密码、设备分类、
        是否有集群属性、是否有专区属性、是否绑定cell、
        带外vlan、带内管理vlan、带内管理接口名称、
        存储vlan、存储网络接口名称、
        业务网络vlan、业务网络接口名称、
        其他网络vlan、其他网络接口名称、
        RAID要求、BIOS配置要求、
        双引擎系统盘分区要求(单位为M)、单引擎系统盘分区要求(单位为M)、
        双引擎推荐操作系统版本、单引擎操作系统、
        存储IP、业务IP、其他IP

        Returns:
            {
                "created": int,
                "updated": int,
                "errors": list
            }

        Raises:
            HostImportError: 导入失败
        """
        import openpyxl
        from io import BytesIO
        try:
            wb = openpyxl.load_workbook(BytesIO(file_obj.read()))
            ws = wb.active
            column_names = [cell.value for cell in ws[1]]

            field_mappings = {
                "序列号": "serial_number",
                "IP地址": "ip_address",
                "主机名": "hostname",
                "ntp地址": "ntp_address",
                "密码": "password",
                "设备分类": "host_type",
                "是否有集群属性": "is_cluster_type",
                "是否有专区属性": "is_zone_type",
                "是否绑定cell": "is_bind_cell_type",
                "带外vlan": "ipmi_vlan",
                "带内管理vlan": "manage_vlan",
                "带内管理接口名称": "manage_nic1",
                "存储vlan": "storage_vlan",
                "存储网络接口名称": "storage_ifname",
                "业务网络vlan": "business_vlan",
                "业务网络接口名称": "business_ifname",
                "其他网络vlan": "other_vlan",
                "其他网络接口名称": "other_ifname",
                "RAID要求": "raid",
                "BIOS配置要求": "bios_config",
                "双引擎系统盘分区要求(单位为M)": "mount_info",
                "单引擎系统盘分区要求(单位为M)": "mount_info",
                "双引擎推荐操作系统版本": "os_version",
                "单引擎操作系统": "os_version",
                "存储IP": "storage_address",
                "业务IP": "business_address",
                "其他IP": "other_address",
            }
            flag_map = {"是": True, "否": False}

            created_count = 0
            updated_count = 0
            errors = []

            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    row_data = dict(zip(column_names, row))
                    host_data = {}
                    for col_name, field_name in field_mappings.items():
                        if col_name in row_data and row_data[col_name] is not None:
                            value = str(row_data[col_name]).strip()
                            if field_name in ("is_cluster_type", "is_zone_type", "is_bind_cell_type"):
                                host_data[field_name] = flag_map.get(value, False)
                            else:
                                host_data[field_name] = value

                    if not host_data.get("hostname") or not host_data.get("ip_address"):
                        continue

                    if not host_data.get("username"):
                        host_data["username"] = "root"

                    existing = Host.objects.filter(hostname=host_data["hostname"]).first()
                    if existing:
                        existing.delete()
                        updated_count += 1
                    else:
                        created_count += 1

                    Host.objects.create(**host_data)
                except Exception as e:
                    errors.append(str(e))

            logger.info(f"Import key cloud completed: created {created_count}, updated {updated_count}, errors {len(errors)}")
            return {
                "created": created_count,
                "updated": updated_count,
                "errors": errors
            }
        except Exception as e:
            logger.error(f"Import key cloud failed: {e}")
            raise HostImportError(str(e))

    @staticmethod
    def export_hosts_to_excel(filters=None):
        """导出主机到 Excel"""
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hosts"
        headers = [
            "hostname", "ip_address", "port", "username", "status",
            "os_type", "host_type", "cluster_name", "serial_number",
            "use_name", "use_for", "manage_nic1", "manage_nic2",
        ]
        ws.append(headers)
        queryset = Host.objects.select_related("cluster").all()
        if filters:
            queryset = queryset.filter(**filters)
        for host in queryset:
            ws.append([
                host.hostname,
                host.ip_address,
                host.port,
                host.username,
                host.status,
                host.os_type,
                host.host_type,
                host.cluster.name if host.cluster else "",
                host.serial_number or "",
                host.use_name or "",
                host.use_for or "",
                host.manage_nic1 or "",
                host.manage_nic2 or "",
            ])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def remote_command(host_id: int, command: str) -> Dict:
        """在远程主机上执行命令

        Returns:
            {
                "stdout": str,
                "stderr": str,
                "exit_code": int
            }

        Raises:
            HostNotFoundError: 主机不存在
            RemoteCommandError: 命令执行失败
        """
        try:
            host = Host.objects.get(pk=host_id)
        except Host.DoesNotExist:
            logger.error(f"Host {host_id} not found")
            raise HostNotFoundError(host_id)

        from backend.utils.ssh import SSHClient
        with SSHClient(host.ip_address, host.port, host.username, host.password, timeout=60) as client:
            stdout, stderr, exit_code = client.execute_command(command)
            if exit_code == 0:
                logger.info(f"Command executed successfully on host {host_id}")
                return {
                    "stdout": stdout,
                    "stderr": stderr,
                    "exit_code": exit_code
                }
            else:
                logger.error(f"Command failed on host {host_id}: {stderr}")
                raise RemoteCommandError(stderr or "命令执行失败")

    @staticmethod
    def batch_update_password(host_ids: List[int], new_password: str = None, key: str = "culinux") -> Dict:
        """批量修改主机密码

        Returns:
            {
                "password": str,
                "updated": int,
                "failed": list
            }
        """
        if not new_password:
            new_password = HostService.generate_random_password()
        encrypted_password = HostService.hash_password(new_password, key)
        updated = 0
        failed = []
        for host_id in host_ids:
            try:
                host = Host.objects.get(pk=host_id)
                host.password = encrypted_password
                host.save()
                updated += 1
            except Host.DoesNotExist:
                failed.append(f"Host {host_id} not found")
            except Exception as e:
                failed.append(f"Host {host_id}: {str(e)}")

        logger.info(f"Batch password update: updated {updated}, failed {len(failed)}")
        return {
            "password": new_password,
            "updated": updated,
            "failed": failed
        }


class VMService:
    """虚拟机服务"""

    @staticmethod
    def list_vms(filters: Optional[dict] = None, page: int = 1, page_size: int = 10):
        """获取VM列表（支持分页和过滤）"""
        queryset = VM.objects.select_related('host', 'cluster').all()
        if filters:
            queryset = queryset.filter(**filters)

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = list(queryset[start:end])

        return {
            'total': total,
            'page': page,
            'page_size': page_size,
            'results': results
        }

    @staticmethod
    def get_vm(vm_id: int) -> Optional[VM]:
        """获取VM详情"""
        try:
            return VM.objects.select_related('host', 'cluster').get(pk=vm_id)
        except VM.DoesNotExist:
            return None

    @staticmethod
    def create_vm(data: dict) -> VM:
        """创建VM"""
        return VM.objects.create(**data)

    @staticmethod
    def update_vm(vm_id: int, data: dict) -> Optional[VM]:
        """更新VM"""
        try:
            vm = VM.objects.get(pk=vm_id)
            for key, value in data.items():
                setattr(vm, key, value)
            vm.save()
            return vm
        except VM.DoesNotExist:
            return None

    @staticmethod
    def delete_vm(vm_id: int) -> bool:
        """删除VM"""
        try:
            vm = VM.objects.get(pk=vm_id)
            vm.delete()
            return True
        except VM.DoesNotExist:
            return False

    @staticmethod
    def start_vm(vm_id: int) -> VMOperationResult:
        """
        启动 VM

        Args:
            vm_id: VM ID

        Returns:
            VMOperationResult: VM 操作结果

        Raises:
            VMNotFoundError: VM不存在
            VMOperationError: 操作失败
        """
        try:
            vm = VM.objects.get(pk=vm_id)
        except VM.DoesNotExist:
            logger.error(f"VM {vm_id} not found")
            raise VMNotFoundError(vm_id)

        try:
            from backend.utils.libvirt_client import LibvirtClient

            host = vm.host
            client = LibvirtClient(
                host=host.ip_address,
                username=host.username,
                password=host.password,
            )
            conn = client._get_conn()
            if conn is None:
                vm.status = 'running'
                vm.save()
                logger.info(f"VM {vm_id} started (simulation mode)")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message='VM启动成功（模拟模式）',
                    is_simulation=True,
                    new_status='running'
                )
            success, message = client.start_domain(vm.name)
            client.close()

            if success:
                vm.status = 'running'
                vm.save()
                logger.info(f"VM {vm_id} started")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message=message,
                    is_simulation=False,
                    new_status='running'
                )
            else:
                logger.error(f"Failed to start VM {vm_id}: {message}")
                raise VMOperationError(message)
        except ImportError:
            vm.status = 'running'
            vm.save()
            logger.info(f"VM {vm_id} started (simulation mode - libvirt not installed)")
            return VMOperationResult(
                success=True,
                vm=vm,
                message='VM启动成功（模拟模式）',
                is_simulation=True,
                new_status='running'
            )
        except Exception as e:
            logger.error(f"Failed to start VM {vm_id}: {e}")
            if 'connection' in str(e).lower() or 'Failed to connect' in str(e):
                vm.status = 'running'
                vm.save()
                logger.info(f"VM {vm_id} started (simulation mode - connection failed)")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message='VM启动成功（模拟模式）',
                    is_simulation=True,
                    new_status='running'
                )
            raise VMOperationError(str(e))

    @staticmethod
    def stop_vm(vm_id: int) -> VMOperationResult:
        """
        停止 VM（强制关机）

        Args:
            vm_id: VM ID

        Returns:
            VMOperationResult: VM 操作结果

        Raises:
            VMNotFoundError: VM不存在
            VMOperationError: 操作失败
        """
        try:
            vm = VM.objects.get(pk=vm_id)
        except VM.DoesNotExist:
            logger.error(f"VM {vm_id} not found")
            raise VMNotFoundError(vm_id)

        try:
            from backend.utils.libvirt_client import LibvirtClient

            host = vm.host
            client = LibvirtClient(
                host=host.ip_address,
                username=host.username,
                password=host.password,
            )
            conn = client._get_conn()
            if conn is None:
                vm.status = 'stopped'
                vm.save()
                logger.info(f"VM {vm_id} stopped (simulation mode)")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message='VM停止成功（模拟模式）',
                    is_simulation=True,
                    new_status='stopped'
                )
            success, message = client.stop_domain(vm.name)
            client.close()

            if success:
                vm.status = 'stopped'
                vm.save()
                logger.info(f"VM {vm_id} stopped")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message=message,
                    is_simulation=False,
                    new_status='stopped'
                )
            else:
                logger.error(f"Failed to stop VM {vm_id}: {message}")
                raise VMOperationError(message)
        except ImportError:
            vm.status = 'stopped'
            vm.save()
            logger.info(f"VM {vm_id} stopped (simulation mode - libvirt not installed)")
            return VMOperationResult(
                success=True,
                vm=vm,
                message='VM停止成功（模拟模式）',
                is_simulation=True,
                new_status='stopped'
            )
        except Exception as e:
            logger.error(f"Failed to stop VM {vm_id}: {e}")
            if 'connection' in str(e).lower() or 'Failed to connect' in str(e):
                vm.status = 'stopped'
                vm.save()
                logger.info(f"VM {vm_id} stopped (simulation mode - connection failed)")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message='VM停止成功（模拟模式）',
                    is_simulation=True,
                    new_status='stopped'
                )
            raise VMOperationError(str(e))

    @staticmethod
    def reboot_vm(vm_id: int) -> VMOperationResult:
        """
        重启 VM

        Args:
            vm_id: VM ID

        Returns:
            VMOperationResult: VM 操作结果

        Raises:
            VMNotFoundError: VM不存在
            VMOperationError: 操作失败
        """
        try:
            vm = VM.objects.get(pk=vm_id)
        except VM.DoesNotExist:
            logger.error(f"VM {vm_id} not found")
            raise VMNotFoundError(vm_id)

        try:
            from backend.utils.libvirt_client import LibvirtClient

            host = vm.host
            client = LibvirtClient(
                host=host.ip_address,
                username=host.username,
                password=host.password,
            )
            conn = client._get_conn()
            if conn is None:
                vm.status = 'running'
                vm.save()
                logger.info(f"VM {vm_id} rebooted (simulation mode)")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message='VM重启成功（模拟模式）',
                    is_simulation=True,
                    new_status='running'
                )
            success, message = client.reboot_domain(vm.name)
            client.close()

            if success:
                vm.status = 'running'
                vm.save()
                logger.info(f"VM {vm_id} rebooted")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message=message,
                    is_simulation=False,
                    new_status='running'
                )
            else:
                logger.error(f"Failed to reboot VM {vm_id}: {message}")
                raise VMOperationError(message)
        except ImportError:
            vm.status = 'running'
            vm.save()
            logger.info(f"VM {vm_id} rebooted (simulation mode - libvirt not installed)")
            return VMOperationResult(
                success=True,
                vm=vm,
                message='VM重启成功（模拟模式）',
                is_simulation=True,
                new_status='running'
            )
        except Exception as e:
            logger.error(f"Failed to reboot VM {vm_id}: {e}")
            if 'connection' in str(e).lower() or 'Failed to connect' in str(e):
                vm.status = 'running'
                vm.save()
                logger.info(f"VM {vm_id} rebooted (simulation mode - connection failed)")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message='VM重启成功（模拟模式）',
                    is_simulation=True,
                    new_status='running'
                )
            raise VMOperationError(str(e))

    @staticmethod
    def pause_vm(vm_id: int) -> VMOperationResult:
        """
        暂停 VM

        Args:
            vm_id: VM ID

        Returns:
            VMOperationResult: VM 操作结果

        Raises:
            VMNotFoundError: VM不存在
            VMOperationError: 操作失败
        """
        try:
            vm = VM.objects.get(pk=vm_id)
        except VM.DoesNotExist:
            logger.error(f"VM {vm_id} not found")
            raise VMNotFoundError(vm_id)

        try:
            from backend.utils.libvirt_client import LibvirtClient

            host = vm.host
            client = LibvirtClient(
                host=host.ip_address,
                username=host.username,
                password=host.password,
            )
            conn = client._get_conn()
            if conn is None:
                vm.status = 'paused'
                vm.save()
                logger.info(f"VM {vm_id} paused (simulation mode)")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message='VM暂停成功（模拟模式）',
                    is_simulation=True,
                    new_status='paused'
                )
            success, message = client.pause_domain(vm.name)
            client.close()

            if success:
                vm.status = 'paused'
                vm.save()
                logger.info(f"VM {vm_id} paused")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message=message,
                    is_simulation=False,
                    new_status='paused'
                )
            else:
                logger.error(f"Failed to pause VM {vm_id}: {message}")
                raise VMOperationError(message)
        except ImportError:
            vm.status = 'paused'
            vm.save()
            logger.info(f"VM {vm_id} paused (simulation mode - libvirt not installed)")
            return VMOperationResult(
                success=True,
                vm=vm,
                message='VM暂停成功（模拟模式）',
                is_simulation=True,
                new_status='paused'
            )
        except Exception as e:
            logger.error(f"Failed to pause VM {vm_id}: {e}")
            if 'connection' in str(e).lower() or 'Failed to connect' in str(e):
                vm.status = 'paused'
                vm.save()
                logger.info(f"VM {vm_id} paused (simulation mode - connection failed)")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message='VM暂停成功（模拟模式）',
                    is_simulation=True,
                    new_status='paused'
                )
            raise VMOperationError(str(e))

    @staticmethod
    def resume_vm(vm_id: int) -> VMOperationResult:
        """
        恢复 VM

        Args:
            vm_id: VM ID

        Returns:
            VMOperationResult: VM 操作结果

        Raises:
            VMNotFoundError: VM不存在
            VMOperationError: 操作失败
        """
        try:
            vm = VM.objects.get(pk=vm_id)
        except VM.DoesNotExist:
            logger.error(f"VM {vm_id} not found")
            raise VMNotFoundError(vm_id)

        try:
            from backend.utils.libvirt_client import LibvirtClient

            host = vm.host
            client = LibvirtClient(
                host=host.ip_address,
                username=host.username,
                password=host.password,
            )
            conn = client._get_conn()
            if conn is None:
                vm.status = 'running'
                vm.save()
                logger.info(f"VM {vm_id} resumed (simulation mode)")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message='VM恢复成功（模拟模式）',
                    is_simulation=True,
                    new_status='running'
                )
            success, message = client.resume_domain(vm.name)
            client.close()

            if success:
                vm.status = 'running'
                vm.save()
                logger.info(f"VM {vm_id} resumed")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message=message,
                    is_simulation=False,
                    new_status='running'
                )
            else:
                logger.error(f"Failed to resume VM {vm_id}: {message}")
                raise VMOperationError(message)
        except ImportError:
            vm.status = 'running'
            vm.save()
            logger.info(f"VM {vm_id} resumed (simulation mode - libvirt not installed)")
            return VMOperationResult(
                success=True,
                vm=vm,
                message='VM恢复成功（模拟模式）',
                is_simulation=True,
                new_status='running'
            )
        except Exception as e:
            logger.error(f"Failed to resume VM {vm_id}: {e}")
            if 'connection' in str(e).lower() or 'Failed to connect' in str(e):
                vm.status = 'running'
                vm.save()
                logger.info(f"VM {vm_id} resumed (simulation mode - connection failed)")
                return VMOperationResult(
                    success=True,
                    vm=vm,
                    message='VM恢复成功（模拟模式）',
                    is_simulation=True,
                    new_status='running'
                )
            raise VMOperationError(str(e))

    @staticmethod
    def get_vm_status(vm_id: int) -> str:
        """
        获取 VM 状态

        Args:
            vm_id: VM ID

        Returns:
            VM 状态字符串

        Raises:
            VMNotFoundError: VM不存在
        """
        try:
            vm = VM.objects.get(pk=vm_id)
        except VM.DoesNotExist:
            logger.error(f"VM {vm_id} not found")
            raise VMNotFoundError(vm_id)

        try:
            from backend.utils.libvirt_client import LibvirtClient

            host = vm.host
            client = LibvirtClient(
                host=host.ip_address,
                username=host.username,
                password=host.password,
            )
            state = client.get_domain_state(vm.name)
            client.close()

            if state is not None:
                libvirt_status = LIBVIRT_STATE_MAP.get(state, 'unknown')
                logger.info(f"Got VM {vm_id} status from libvirt: {libvirt_status}")
                return libvirt_status
            else:
                logger.info(f"Got VM {vm_id} status from database: {vm.status}")
                return vm.status
        except ImportError:
            logger.info(f"Got VM {vm_id} status from database (libvirt not installed): {vm.status}")
            return vm.status
        except Exception as e:
            logger.error(f"Failed to get VM status {vm_id} from libvirt, using database: {e}")
            return vm.status

    @staticmethod
    def delete_vm_from_libvirt(vm_id: int) -> VMOperationResult:
        """
        从 libvirt 删除 VM（先关机再删除定义）

        Args:
            vm_id: VM ID

        Returns:
            VMOperationResult: VM 操作结果

        Raises:
            VMNotFoundError: VM不存在
            VMOperationError: 操作失败
        """
        try:
            vm = VM.objects.get(pk=vm_id)
        except VM.DoesNotExist:
            logger.error(f"VM {vm_id} not found")
            raise VMNotFoundError(vm_id)

        try:
            from backend.utils.libvirt_client import LibvirtClient

            host = vm.host
            client = LibvirtClient(
                host=host.ip_address,
                username=host.username,
                password=host.password,
            )
            conn = client._get_conn()
            if conn is None:
                vm.delete()
                logger.info(f"VM {vm_id} deleted (simulation mode)")
                return VMOperationResult(
                    success=True,
                    vm=None,
                    message='VM删除成功（模拟模式）',
                    is_simulation=True,
                    new_status=None
                )

            client.stop_domain(vm.name)
            success, message = client.undefine_domain(vm.name)
            client.close()

            if success:
                vm.delete()
                logger.info(f"VM {vm_id} deleted from libvirt and database")
                return VMOperationResult(
                    success=True,
                    vm=None,
                    message=message,
                    is_simulation=False,
                    new_status=None
                )
            else:
                logger.error(f"Failed to delete VM {vm_id} from libvirt: {message}")
                raise VMOperationError(message)
        except ImportError:
            vm.delete()
            logger.info(f"VM {vm_id} deleted (simulation mode - libvirt not installed)")
            return VMOperationResult(
                success=True,
                vm=None,
                message='VM删除成功（模拟模式）',
                is_simulation=True,
                new_status=None
            )
        except Exception as e:
            logger.error(f"Failed to delete VM {vm_id}: {e}")
            if 'connection' in str(e).lower() or 'Failed to connect' in str(e):
                vm.delete()
                logger.info(f"VM {vm_id} deleted (simulation mode - connection failed)")
                return VMOperationResult(
                    success=True,
                    vm=None,
                    message='VM删除成功（模拟模式）',
                    is_simulation=True,
                    new_status=None
                )
            raise VMOperationError(str(e))

    @staticmethod
    def create_vm_in_libvirt(vm_id: int) -> Dict:
        """
        在 libvirt 中创建 VM

        Args:
            vm_id: VM ID

        Returns:
            {'message': str}
        """
        try:
            vm = VM.objects.get(pk=vm_id)
        except VM.DoesNotExist:
            raise VMNotFoundError(vm_id)

        try:
            from backend.utils.libvirt_client import LibvirtClient

            host = vm.host
            client = LibvirtClient(
                host=host.ip_address,
                username=host.username,
                password=host.password,
            )
            conn = client._get_conn()
            if conn is None:
                raise VMOperationError('libvirt 连接失败，无法创建 VM')

            # 生成 Domain XML
            xml = VMService._generate_domain_xml(vm)
            success, message = client.create_domain(xml)
            client.close()

            if success:
                vm.status = 'running'
                vm.save()
            else:
                raise VMOperationError(message)
            return {'message': message}
        except ImportError:
            raise VMOperationError('libvirt 未安装')
        except VMOperationError:
            raise
        except Exception as e:
            logger.error(f"Failed to create VM {vm_id}: {e}")
            raise VMOperationError(str(e))

    @staticmethod
    def _generate_domain_xml(vm: VM) -> str:
        """
        生成 libvirt Domain XML

        Args:
            vm: VM 实例

        Returns:
            XML 字符串
        """
        # 获取数据盘信息
        datadisks_xml = ""
        if vm.datadisk and isinstance(vm.datadisk, list):
            for i, disk in enumerate(vm.datadisk):
                disk_path = disk.get('path', '')
                if disk_path:
                    datadisks_xml += f'''
    <disk type='file' device='disk'>
      <driver name='qemu' type='raw'/>
      <source file='{disk_path}'/>
      <target dev='vd{"bcdefghijklmnopqrstuvwxyz"[i]}' bus='virtio'/>
    </disk>'''

        # 获取网桥信息
        netdevname = vm.vm_network_bridge or "mgmt"
        network_xml = f'''
    <interface type='bridge'>
      <source bridge='{netdevname}'/>
      <model type='virtio'/>
    </interface>'''

        # 内存单位转换（字节 -> GiB）
        memory_gib = max(vm.memory // (1024**3), 1)

        xml = f'''<domain type='kvm'>
  <name>{vm.name}</name>
  <memory unit='GiB'>{memory_gib}</memory>
  <currentMemory unit='GiB'>{memory_gib}</currentMemory>
  <vcpu placement='static'>{vm.vcpu}</vcpu>
  <os>
    <type arch='x86_64' machine='pc-i440fx-2.9'>hvm</type>
    <boot dev='hd'/>
  </os>
  <devices>
    <disk type='file' device='disk'>
      <driver name='qemu' type='qcow2'/>
      <source file='{vm.vm_image_path}'/>
      <target dev='vda' bus='virtio'/>
    </disk>
    {datadisks_xml}
    {network_xml}
    <graphics type='vnc' port='-1' autoport='yes' listen='0.0.0.0'>
      <listen type='address' address='0.0.0.0'/>
    </graphics>
    <channel type='unix'>
      <target type='virtio' name='org.qemu.guest_agent.0'/>
      <address type='virtio-serial' controller='0' bus='0' port='1'/>
    </channel>
  </devices>
</domain>'''
        return xml
