"""用户相关服务"""
from typing import Optional
from django.db import transaction
from backend.models import Users, UserAuthority, Authority


class UserService:
    """用户服务"""

    @staticmethod
    def get_user(user_id: int) -> Optional[Users]:
        """获取用户详情"""
        try:
            return Users.objects.get(pk=user_id)
        except Users.DoesNotExist:
            return None

    @staticmethod
    def create_user(data: dict) -> Users:
        """创建用户"""
        return Users.objects.create(**data)

    @staticmethod
    def update_user(user_id: int, data: dict) -> Optional[Users]:
        """更新用户"""
        try:
            user = Users.objects.get(pk=user_id)
            for key, value in data.items():
                setattr(user, key, value)
            user.save()
            return user
        except Users.DoesNotExist:
            return None

    @staticmethod
    def delete_user(user_id: int) -> bool:
        """删除用户"""
        try:
            user = Users.objects.get(pk=user_id)
            user.delete()
            return True
        except Users.DoesNotExist:
            return False

    @staticmethod
    def list_users(filters: Optional[dict] = None, page: int = 1, page_size: int = 10):
        """获取用户列表（支持分页和过滤）"""
        queryset = Users.objects.all()
        if filters:
            queryset = queryset.filter(**filters)

        total = queryset.count()
        start = (page - 1) * page_size
        end = start + page_size
        results = list(queryset[start:end])

        return {
            'total': total,
            'page': page,
            'page_size': page_size,
            'results': results
        }

    @staticmethod
    def import_users_from_excel(file_obj):
        """从 Excel 批量导入用户"""
        import openpyxl
        from io import BytesIO
        try:
            wb = openpyxl.load_workbook(BytesIO(file_obj.read()))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            created_count = 0
            updated_count = 0
            errors = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    data = dict(zip(headers, row))
                    username = data.get("user") or data.get("username")
                    if not username:
                        continue
                    user, created = Users.objects.update_or_create(
                        user=username,
                        defaults={
                            "nickname": data.get("nickname", username),
                            "phone": str(data.get("phone", "")),
                            "email": data.get("email", ""),
                            "enable": int(data.get("enable", 1)),
                            "theme": data.get("theme", "light"),
                        }
                    )
                    # 如果有密码则设置
                    password = data.get("password")
                    if password:
                        user.set_password(str(password))
                        user.save()
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                except Exception as e:
                    errors.append(str(e))
            return {
                "success": True,
                "created": created_count,
                "updated": updated_count,
                "errors": errors,
                "message": f"导入完成：新建 {created_count} 条，更新 {updated_count} 条，错误 {len(errors)} 条",
            }
        except Exception as e:
            return {"success": False, "message": str(e), "created": 0, "updated": 0, "errors": [str(e)]}

    @staticmethod
    def export_users_to_excel():
        """导出用户到 Excel"""
        import openpyxl
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        headers = ["user", "nickname", "phone", "email", "enable", "theme", "avatar"]
        ws.append(headers)
        for user in Users.objects.all():
            ws.append([
                user.user,
                user.nickname,
                user.phone,
                user.email,
                user.enable,
                user.theme,
                user.avatar,
            ])
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


class UserAuthorityService:
    """用户角色关联服务"""

    @staticmethod
    def get_user_authorities(user_id: int):
        """获取用户的所有角色"""
        return Authority.objects.filter(userauthority__user_id=user_id)

    @staticmethod
    def set_user_roles(user_id: int, role_ids: list) -> bool:
        """设置用户角色（覆盖式）"""
        try:
            user = Users.objects.get(pk=user_id)
            with transaction.atomic():
                # 先删除所有角色关联
                UserAuthority.objects.filter(user=user).delete()
                # 再创建新关联
                for role_id in role_ids:
                    UserAuthority.objects.create(user=user, authority_id=role_id)
            return True
        except Users.DoesNotExist:
            return False

    @staticmethod
    def add_user_authority(user_id: int, authority_id: int) -> Optional[UserAuthority]:
        """为用户添加角色"""
        try:
            user = Users.objects.get(pk=user_id)
            authority = Authority.objects.get(pk=authority_id)
            return UserAuthority.objects.get_or_create(user=user, authority=authority)
        except (Users.DoesNotExist, Authority.DoesNotExist):
            return None

    @staticmethod
    def remove_user_authority(user_id: int, authority_id: int) -> bool:
        """移除用户角色"""
        try:
            ua = UserAuthority.objects.get(user_id=user_id, authority_id=authority_id)
            ua.delete()
            return True
        except UserAuthority.DoesNotExist:
            return False
